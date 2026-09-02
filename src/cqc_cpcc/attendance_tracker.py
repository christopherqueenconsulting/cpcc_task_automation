"""Sync withdrawal records to the online Attendance Tracker.

The tracker is a Microsoft/SharePoint-hosted Excel workbook, so this reuses the
project's existing Microsoft + Duo login path.

Two rules govern everything here:

1. **Dry run by default.** A real write happens only when the caller explicitly
   asks for one. A dry run navigates, reads, and reports exactly what it would
   append, and changes nothing.
2. **Never append after a failed read.** De-duplication depends on knowing what
   the sheet already contains. If that read fails, the sync aborts rather than
   appending blind and duplicating the tracker.
"""

import base64
import io
import re
import time
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from urllib.parse import urlparse

from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC

from cqc_cpcc.utilities.logger import logger
from cqc_cpcc.utilities.selenium_util import close_tab
from cqc_cpcc.utilities.utils import login_if_needed
from cqc_cpcc.withdrawals import CSV_FIELDNAMES, WithdrawalRecord, record_key

# Hosts that indicate a Microsoft-hosted workbook.
_SHAREPOINT_HOST_MARKERS = (
    "sharepoint.com",
    "onedrive.live.com",
    "office.com",
    "officeapps.live.com",
)

# The workbook download can be slow on a cold Excel Online session.
WORKBOOK_FETCH_TIMEOUT_SECONDS = 120

# Excel Online is a remote renderer; typing needs pacing to stay in sync.
CELL_ENTRY_PAUSE_SECONDS = 1.0
AUTOSAVE_SETTLE_SECONDS = 4.0

# Excel Online renders its grid inside a nested web-application iframe.
_EXCEL_FRAME_SELECTORS = (
    "iframe.WACFrame",
    "iframe[name='WacFrame_Excel_0']",
    "iframe#WebApplicationFrame",
    "iframe[title*='Excel']",
)


class TrackerSyncError(Exception):
    """Raised when the tracker cannot be read or written safely."""


@dataclass
class SyncResult:
    """What a sync did, or would have done."""

    tracker_url: str = ""
    dry_run: bool = True
    existing_count: int = 0
    to_append: list = field(default_factory=list)
    appended: int = 0
    skipped_existing: int = 0

    def describe(self) -> str:
        verb = "Would append" if self.dry_run else "Appended"
        return (
            "Tracker sync (%s): %s existing row(s) read, %s already present, "
            "%s %s row(s)."
            % (
                "dry run" if self.dry_run else "live",
                self.existing_count,
                self.skipped_existing,
                verb,
                len(self.to_append),
            )
        )


def open_attendance_tracker(driver, wait, attendance_tracker_url: str) -> str:
    """Open the tracker in a new tab and complete login. Returns the tab handle."""
    handles = set(driver.window_handles)

    driver.switch_to.new_window('tab')
    wait.until(EC.new_window_is_opened(handles))
    tracker_tab = driver.current_window_handle

    driver.get(attendance_tracker_url)
    logger.info("Navigated to the Attendance Tracker: %s", attendance_tracker_url)

    # The tracker is behind the same Microsoft SSO (and Duo MFA) as everything else.
    login_if_needed(driver)

    return tracker_tab


class TrackerAdapter(ABC):
    """How to read from and write to one flavour of online tracker."""

    _tracker_url: str = ""
    _next_free_row: int | None = None

    def bind(self, tracker_url: str) -> "TrackerAdapter":
        """Tell the adapter which workbook it is operating on."""
        self._tracker_url = tracker_url
        return self

    @abstractmethod
    def read_existing_keys(self, driver, wait) -> set:
        """Return the ``(course, student id)`` keys already on the tracker."""

    @abstractmethod
    def append_records(self, driver, wait, records: list) -> int:
        """Append records to the tracker and return how many were written."""


class SharePointExcelAdapter(TrackerAdapter):
    """Excel Online (SharePoint / OneDrive for Business).

    **Reads do not scrape the page.** Verified live 2026-09-01: the Excel Online
    grid is rendered to ``<canvas>`` and exposes no ARIA rows or gridcells at all
    (``[role=row]`` count is zero after a full minute of loading), so any
    DOM-scraping read is impossible, not merely fragile.

    Instead the workbook is downloaded through ``download.aspx`` using a ``fetch``
    issued *from inside the page*, so the browser's own session cookies apply, and
    parsed with openpyxl. That yields exact cell values -- which is what the
    duplicate check needs, since appending after a bad read would double the
    tracker.
    """

    # Header row, then data. Matches the live "Instructor Inputs" sheet.
    SHEET_NAME = "Instructor Inputs"
    COURSE_HEADER = "Course and Section"
    STUDENT_ID_HEADER = "Student ID"

    # Runs in the page origin; SharePoint rejects cookies lifted out to a separate
    # HTTP client (verified: 401).
    _FETCH_WORKBOOK_JS = """
        var url = arguments[0], done = arguments[arguments.length - 1];
        fetch(url, {credentials: 'include'})
          .then(function (response) {
            if (!response.ok) { done({status: response.status}); return; }
            return response.arrayBuffer().then(function (buffer) {
              var bytes = new Uint8Array(buffer), chunk = 0x8000, parts = [];
              for (var i = 0; i < bytes.length; i += chunk) {
                var slice = bytes.subarray(i, i + chunk);
                parts.push(String.fromCharCode.apply(null, slice));
              }
              done({status: response.status, b64: btoa(parts.join('')),
                    size: bytes.length});
            });
          })
          .catch(function (error) { done({error: String(error).slice(0, 200)}); });
    """

    @staticmethod
    def download_url(tracker_url: str) -> str | None:
        """Build the workbook download URL from a SharePoint share link."""
        guid = re.search(r"sourcedoc=%7B([0-9A-Fa-f-]+)%7D", tracker_url or "")
        site = re.match(
            r"(https://[^/]+)(/personal/[^/]+|/sites/[^/]+)/",
            (tracker_url or "").replace("/:x:/r", ""),
        )
        if not guid or not site:
            return None
        return "%s%s/_layouts/15/download.aspx?UniqueId=%%7B%s%%7D" % (
            site.group(1), site.group(2), guid.group(1)
        )

    def fetch_workbook(self, driver, tracker_url: str) -> bytes:
        """Download the workbook bytes through the authenticated browser session."""
        url = self.download_url(tracker_url)
        if not url:
            raise TrackerSyncError(
                "Could not derive a download URL from the tracker link. Expected a "
                "SharePoint share URL containing sourcedoc={GUID}."
            )

        driver.switch_to.default_content()
        try:
            driver.set_script_timeout(WORKBOOK_FETCH_TIMEOUT_SECONDS)
            result = driver.execute_async_script(self._FETCH_WORKBOOK_JS, url) or {}
        except Exception as fetch_error:
            raise TrackerSyncError(
                "Workbook download failed: %s" % fetch_error
            ) from fetch_error

        if result.get("error"):
            raise TrackerSyncError("Workbook download failed: %s" % result["error"])
        if not result.get("b64"):
            raise TrackerSyncError(
                "Workbook download returned HTTP %s with no content. The session may "
                "not be signed in to SharePoint." % result.get("status")
            )

        data = base64.b64decode(result["b64"])
        if data[:2] != b"PK":
            raise TrackerSyncError(
                "Downloaded content is not an .xlsx file (got %d bytes). The link may "
                "point at a sign-in page rather than the workbook." % len(data)
            )

        logger.info("Downloaded the tracker workbook (%d bytes).", len(data))
        return data

    def load_sheet(self, workbook_bytes: bytes):
        """Return (header list, data rows) from the instructor-inputs sheet."""
        import openpyxl

        workbook = openpyxl.load_workbook(
            io.BytesIO(workbook_bytes), read_only=True, data_only=True
        )
        sheet_name = (
            self.SHEET_NAME
            if self.SHEET_NAME in workbook.sheetnames
            else workbook.sheetnames[0]
        )
        sheet = workbook[sheet_name]

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise TrackerSyncError(
                "Tracker sheet %r is empty (no header row)." % sheet_name
            )

        header = ["" if cell is None else str(cell).strip() for cell in rows[0]]
        data = [row for row in rows[1:] if any(cell not in (None, "") for cell in row)]
        return header, data

    def read_existing_keys(self, driver, wait) -> set:
        header, data = self.load_sheet(self.fetch_workbook(driver, self._tracker_url))

        try:
            course_index = header.index(self.COURSE_HEADER)
            student_index = header.index(self.STUDENT_ID_HEADER)
        except ValueError:
            raise TrackerSyncError(
                "Could not find the %r and %r columns in the tracker header: %r. "
                "Refusing to append without a reliable duplicate check."
                % (self.COURSE_HEADER, self.STUDENT_ID_HEADER, header)
            ) from None

        keys = set()
        for row in data:
            if len(row) <= max(course_index, student_index):
                continue
            course = str(row[course_index] or "").strip().upper()
            # str() not int(): student ids are identifiers, and leading zeros matter.
            student_id = str(row[student_index] or "").strip()
            if student_id.endswith(".0"):
                # openpyxl hands back numeric cells as floats.
                student_id = student_id[:-2]
            if course and student_id:
                keys.add((course, student_id))

        logger.info(
            "Read %s existing row(s) from the tracker (%s data row(s) on the sheet).",
            len(keys), len(data),
        )
        self._next_free_row = len(data) + 2  # +1 header, +1 for one-based rows
        return keys

    # The Name Box's aria-label is wrapped in Unicode LTR marks (\u200e), so an
    # exact-match selector silently finds nothing. Substring matching is required.
    NAME_BOX_SELECTORS = (
        "input[aria-label*='Name Box']",
        "#m_excelWebRenderer_ewaCtl_nameBox",
    )

    @staticmethod
    def _sanitize_cell(value) -> str:
        """Make a value safe to type into a spreadsheet cell.

        Tabs and newlines would be read as cell/row separators and shift every
        following value into the wrong column. A leading =, +, - or @ makes Excel
        treat student-derived text as a formula, so it is prefixed with an
        apostrophe -- Excel's own "this is text" marker, which it does not display.
        """
        text = "" if value is None else str(value)
        text = text.replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()
        if text[:1] in ("=", "+", "-", "@"):
            text = "'" + text
        return text

    def _enter_grid_frame(self, driver) -> None:
        driver.switch_to.default_content()
        for selector in _EXCEL_FRAME_SELECTORS:
            frames = driver.find_elements(By.CSS_SELECTOR, selector)
            if frames:
                driver.switch_to.frame(frames[0])
                return
        raise TrackerSyncError(
            "Could not find the Excel Online web-application frame. The workbook may "
            "not have finished loading."
        )

    def _goto_cell(self, driver, address: str) -> None:
        """Jump to a cell by address using the Name Box."""
        for selector in self.NAME_BOX_SELECTORS:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if not elements:
                continue
            name_box = elements[0]
            name_box.click()
            name_box.send_keys(Keys.CONTROL, "a")
            name_box.send_keys(address)
            name_box.send_keys(Keys.ENTER)
            time.sleep(CELL_ENTRY_PAUSE_SECONDS)
            return

        raise TrackerSyncError(
            "Could not find the Excel Name Box, so the target cell cannot be selected "
            "safely. Refusing to type blind."
        )

    def append_records(self, driver, wait, records: list) -> int:
        """Append rows at the first free row, one row at a time.

        Only columns A-K of *new* rows are touched. The target cell is re-selected
        by address for every row rather than relying on where Excel leaves the
        cursor, so a mis-timed keystroke cannot wander into an existing row or into
        the Navigator's columns L and M.
        """
        if self._next_free_row is None:
            raise TrackerSyncError(
                "Refusing to append: the sheet has not been read, so the first free "
                "row is unknown."
            )

        self._enter_grid_frame(driver)
        written = 0

        try:
            for offset, record in enumerate(records):
                target_row = self._next_free_row + offset
                self._goto_cell(driver, "A%d" % target_row)

                row = record.to_csv_row()
                values = [self._sanitize_cell(row[column]) for column in CSV_FIELDNAMES]

                # One send_keys per row: tabs advance the cell, Enter commits.
                ActionChains(driver).send_keys("\t".join(values)).send_keys(Keys.ENTER).perform()
                time.sleep(CELL_ENTRY_PAUSE_SECONDS)
                written += 1
                logger.info(
                    "Appended row %s: %s / %s",
                    target_row, record.course_and_section, record.student_id,
                )
        finally:
            driver.switch_to.default_content()

        # Excel Online autosaves; give it a moment before anything re-reads.
        time.sleep(AUTOSAVE_SETTLE_SECONDS)
        return written


def build_adapter(tracker_url: str) -> TrackerAdapter:
    """Pick the adapter for a tracker URL based on its host."""
    host = (urlparse(tracker_url or "").hostname or "").lower()

    if any(marker in host for marker in _SHAREPOINT_HOST_MARKERS):
        return SharePointExcelAdapter()

    raise TrackerSyncError(
        "Unsupported Attendance Tracker host: %r. This build supports "
        "Microsoft/SharePoint hosted Excel workbooks. Update "
        "ATTENDANCE_TRACKER_URL, or add an adapter for this host."
        % (host or tracker_url)
    )


def sync_records_to_tracker(
        driver,
        wait,
        tracker_url: str,
        records: list,
        dry_run: bool = True,
        adapter: TrackerAdapter = None,
) -> SyncResult:
    """Append any records missing from the online tracker.

    De-duplication is done against a fresh read of the tracker itself, independent
    of the local CSV, so rows entered by hand online are never re-added.
    """
    result = SyncResult(tracker_url=tracker_url, dry_run=dry_run)

    if not tracker_url:
        raise TrackerSyncError("No Attendance Tracker URL configured.")

    if not records:
        logger.info("No withdrawal records to sync to the tracker.")
        return result

    adapter = (adapter or build_adapter(tracker_url))
    if hasattr(adapter, 'bind'):
        adapter.bind(tracker_url)

    original_tab = driver.current_window_handle
    tracker_tab = None

    try:
        tracker_tab = open_attendance_tracker(driver, wait, tracker_url)

        existing_keys = adapter.read_existing_keys(driver, wait)
        result.existing_count = len(existing_keys)

        for record in records:
            if record_key(record) in existing_keys:
                result.skipped_existing += 1
            else:
                result.to_append.append(record)

        _log_pending_rows(result.to_append)

        if dry_run:
            logger.info("Dry run: the tracker was not modified.")
        elif result.to_append:
            result.appended = adapter.append_records(driver, wait, result.to_append)

        logger.info(result.describe())
        return result

    finally:
        driver.switch_to.default_content()
        if tracker_tab and tracker_tab in driver.window_handles:
            driver.switch_to.window(tracker_tab)
            close_tab(driver)
        if original_tab in driver.window_handles:
            driver.switch_to.window(original_tab)


def _log_pending_rows(records: list) -> None:
    """Log the rows destined for the tracker, in tracker column order."""
    if not records:
        logger.info(
            "Every withdrawal record is already on the tracker. Nothing to append."
        )
        return

    logger.info("Rows for the Attendance Tracker:")
    logger.info(",".join(CSV_FIELDNAMES))
    for record in records:
        row = record.to_csv_row()
        logger.info(",".join(str(row[column]) for column in CSV_FIELDNAMES))


def records_needing_sync(records: list, existing_keys: set) -> list:
    """Pure helper: the records whose keys are not already present."""
    return [record for record in records if record_key(record) not in existing_keys]


__all__ = [
    "SharePointExcelAdapter",
    "SyncResult",
    "TrackerAdapter",
    "TrackerSyncError",
    "WithdrawalRecord",
    "build_adapter",
    "open_attendance_tracker",
    "records_needing_sync",
    "sync_records_to_tracker",
]
