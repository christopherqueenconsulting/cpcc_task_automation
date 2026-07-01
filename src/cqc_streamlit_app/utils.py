#  Copyright (c) 2024. Christopher Queen Consulting LLC (http://www.ChristopherQueenConsulting.com/)
import os
import tempfile
import zipfile
from random import randint
from typing import Any, Optional, Union

import pandas as pd
import streamlit as st
from cqc_cpcc.utilities.file_url_utils import (
    download_file_from_url,
    parse_google_drive_url,
    sanitize_zip_filename,
)
from cqc_cpcc.utilities.language_utils import get_language_from_file_path
from cqc_cpcc.utilities.logger import logger
from langchain_openai import ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit.delta_generator import DeltaGenerator
from streamlit.runtime.uploaded_file_manager import UploadedFile

EXTENSION_TO_LANGUAGES = {
    # Python
    "py": ["python"],
    "pyw": ["python"],

    # C / C++
    "c": ["c"],
    "h": ["c", "cpp"],
    "cpp": ["cpp"],
    "cc": ["cpp"],
    "cxx": ["cpp"],
    "hpp": ["cpp"],
    "hh": ["cpp"],

    # Java
    "java": ["java"],
    "class": ["java"],

    # JavaScript / TypeScript
    "js": ["javascript"],
    "mjs": ["javascript"],
    "cjs": ["javascript"],
    "jsx": ["javascript", "react"],
    "ts": ["typescript"],
    "tsx": ["typescript", "react"],

    # Web / Markup
    "html": ["html"],
    "htm": ["html"],
    "xml": ["xml"],
    "svg": ["xml"],
    "css": ["css"],
    "scss": ["scss"],
    "sass": ["sass"],
    "less": ["less"],

    # Shell
    "sh": ["bash"],
    "bash": ["bash"],
    "zsh": ["bash"],
    "fish": ["bash"],

    # Data / Config
    "json": ["json"],
    "json5": ["json"],
    "yaml": ["yaml"],
    "yml": ["yaml"],
    "toml": ["toml"],
    "ini": ["ini"],
    "env": ["dotenv"],
    "conf": ["apacheconf", "nginx"],

    # Markdown / Text
    "md": ["markdown"],
    "markdown": ["markdown"],
    "txt": ["text"],

    # SQL
    "sql": ["sql"],
    "psql": ["sql"],

    # Go
    "go": ["go"],
    "mod": ["go"],

    # Rust
    "rs": ["rust"],

    # Ruby
    "rb": ["ruby"],

    # PHP
    "php": ["php"],

    # Swift
    "swift": ["swift"],

    # Kotlin
    "kt": ["kotlin"],
    "kts": ["kotlin"],

    # C#
    "cs": ["csharp"],

    # Objective-C / MATLAB conflict
    "m": ["objective-c", "matlab"],
    "mm": ["objective-c"],

    # Dart
    "dart": ["dart"],

    # R
    "r": ["r"],
    "R": ["r"],

    # Scala
    "scala": ["scala"],

    # Groovy
    "groovy": ["groovy"],

    # Haskell
    "hs": ["haskell"],

    # Lua
    "lua": ["lua"],

    # Perl
    "pl": ["perl", "prolog"],
    "pm": ["perl"],

    # PowerShell
    "ps1": ["powershell"],

    # Windows scripts
    "bat": ["batch"],
    "cmd": ["batch"],

    # Assembly
    "asm": ["assembly"],
    "s": ["assembly"],

    # Docker
    "dockerfile": ["docker"],
    "docker": ["docker"],

    # Make
    "makefile": ["makefile"],
    "mk": ["makefile"],

    # GraphQL
    "graphql": ["graphql"],
    "gql": ["graphql"],

    # Protobuf
    "proto": ["protobuf"],

    # LaTeX
    "tex": ["latex"],

    # Terraform / HCL
    "tf": ["hcl"],
    "tfvars": ["hcl"],

    # Vim
    "vim": ["vim"],

    # Zig
    "zig": ["zig"],

    # WebAssembly
    "wasm": ["wasm"],

    # Misc
    "log": ["log"],
}

mime_types_str = """
.3dm	x-world/x-3dmf
.3dmf	x-world/x-3dmf
.7z	application/x-7z-compressed
.a	application/octet-stream
.aab	application/x-authorware-bin
.aam	application/x-authorware-map
.aas	application/x-authorware-seg
.abc	text/vnd.abc
.acgi	text/html
.afl	video/animaflex
.ai	application/postscript
.aif	audio/aiff
.aif	audio/x-aiff
.aifc	audio/aiff
.aifc	audio/x-aiff
.aiff	audio/aiff
.aiff	audio/x-aiff
.aim	application/x-aim
.aip	text/x-audiosoft-intra
.ani	application/x-navi-animation
.aos	application/x-nokia-9000-communicator-add-on-software
.aps	application/mime
.arc	application/octet-stream
.arj	application/arj
.arj	application/octet-stream
.art	image/x-jg
.asf	video/x-ms-asf
.asm	text/x-asm
.asp	text/asp
.asx	application/x-mplayer2
.asx	video/x-ms-asf
.asx	video/x-ms-asf-plugin
.au	audio/basic
.au	audio/x-au
.avi	application/x-troff-msvideo
.avi	video/avi
.avi	video/msvideo
.avi	video/x-msvideo
.avs	video/avs-video
.bcpio	application/x-bcpio
.bin	application/mac-binary
.bin	application/macbinary
.bin	application/octet-stream
.bin	application/x-binary
.bin	application/x-macbinary
.bm	image/bmp
.bmp	image/bmp
.bmp	image/x-windows-bmp
.boo	application/book
.book	application/book
.boz	application/x-bzip2
.bsh	application/x-bsh
.bz	application/x-bzip
.bz2	application/x-bzip2
.c	text/plain
.c	text/x-c
.c++	text/plain
.cat	application/vnd.ms-pki.seccat
.cc	text/plain
.cc	text/x-c
.ccad	application/clariscad
.cco	application/x-cocoa
.cdf	application/cdf
.cdf	application/x-cdf
.cdf	application/x-netcdf
.cer	application/pkix-cert
.cer	application/x-x509-ca-cert
.cha	application/x-chat
.chat	application/x-chat
.class	application/java
.class	application/java-byte-code
.class	application/x-java-class
.com	application/octet-stream
.com	text/plain
.conf	text/plain
.cpio	application/x-cpio
.cpp	text/x-c
.cpt	application/mac-compactpro
.cpt	application/x-compactpro
.cpt	application/x-cpt
.crl	application/pkcs-crl
.crl	application/pkix-crl
.crt	application/pkix-cert
.crt	application/x-x509-ca-cert
.crt	application/x-x509-user-cert
.csh	application/x-csh
.csh	text/x-script.csh
.css	application/x-pointplus
.css	text/css
.csv	text/csv
.cxx	text/plain
.dcr	application/x-director
.deepv	application/x-deepv
.def	text/plain
.der	application/x-x509-ca-cert
.dif	video/x-dv
.dir	application/x-director
.dl	video/dl
.dl	video/x-dl
.doc	application/msword
.docx	application/vnd.openxmlformats-officedocument.wordprocessingml.document
.dot	application/msword
.dp	application/commonground
.drw	application/drafting
.dump	application/octet-stream
.dv	video/x-dv
.dvi	application/x-dvi
.dwf	model/vnd.dwf
.dwg	application/acad
.dwg	image/vnd.dwg
.dwg	image/x-dwg
.dxf	application/dxf
.dxf	image/vnd.dwg
.dxf	image/x-dwg
.dxr	application/x-director
.el	text/x-script.elisp
.elc	application/x-elc
.env	application/x-envoy
.eot	application/vnd.ms-fontobject
.eps	application/postscript
.es	application/x-esrehber
.etx	text/x-setext
.evy	application/envoy
.evy	application/x-envoy
.exe	application/octet-stream
.f	text/plain
.f	text/x-fortran
.f77	text/x-fortran
.f90	text/plain
.f90	text/x-fortran
.fdf	application/vnd.fdf
.fif	application/fractals
.fif	image/fif
.flac	audio/flac
.fli	video/fli
.fli	video/x-fli
.flo	image/florian
.flx	text/vnd.fmi.flexstor
.fmf	video/x-atomic3d-feature
.for	text/plain
.for	text/x-fortran
.fpx	image/vnd.fpx
.fpx	image/vnd.net-fpx
.frl	application/freeloader
.funk	audio/make
.g	text/plain
.g3	image/g3fax
.gif	image/gif
.gl	video/gl
.gl	video/x-gl
.gsd	audio/x-gsm
.gsm	audio/x-gsm
.gsp	application/x-gsp
.gss	application/x-gss
.gtar	application/x-gtar
.gz	application/x-compressed
.gz	application/x-gzip
.gzip	application/x-gzip
.gzip	multipart/x-gzip
.h	text/plain
.h	text/x-h
.hdf	application/x-hdf
.help	application/x-helpfile
.hgl	application/vnd.hp-hpgl
.hh	text/plain
.hh	text/x-h
.hlb	text/x-script
.hlp	application/hlp
.hlp	application/x-helpfile
.hlp	application/x-winhelp
.hpg	application/vnd.hp-hpgl
.hpgl	application/vnd.hp-hpgl
.hqx	application/binhex
.hqx	application/binhex4
.hqx	application/mac-binhex
.hqx	application/mac-binhex40
.hqx	application/x-binhex40
.hqx	application/x-mac-binhex40
.hta	application/hta
.htc	text/x-component
.htm	text/html
.html	text/html
.htmls	text/html
.htt	text/webviewhtml
.htx	text/html
.ice	x-conference/x-cooltalk
.ico	image/x-icon
.ics	text/calendar
.idc	text/plain
.ief	image/ief
.iefs	image/ief
.iges	application/iges
.iges	model/iges
.igs	application/iges
.igs	model/iges
.ima	application/x-ima
.imap	application/x-httpd-imap
.inf	application/inf
.ins	application/x-internett-signup
.ip	application/x-ip2
.isu	video/x-isvideo
.it	audio/it
.iv	application/x-inventor
.ivr	i-world/i-vrml
.ivy	application/x-livescreen
.jam	audio/x-jam
.jav	text/plain
.jav	text/x-java-source
.java	text/plain
.java	text/x-java-source
.jcm	application/x-java-commerce
.jfif	image/jpeg
.jfif	image/pjpeg
.jfif-tbnl	image/jpeg
.jpe	image/jpeg
.jpe	image/pjpeg
.jpeg	image/jpeg
.jpeg	image/pjpeg
.jpg	image/jpeg
.jpg	image/pjpeg
.jps	image/x-jps
.js	application/x-javascript
.js	application/javascript
.js	application/ecmascript
.js	text/javascript
.js	text/ecmascript
.json	application/json
.jut	image/jutvision
.kar	audio/midi
.kar	music/x-karaoke
.ksh	application/x-ksh
.ksh	text/x-script.ksh
.la	audio/nspaudio
.la	audio/x-nspaudio
.lam	audio/x-liveaudio
.latex	application/x-latex
.lha	application/lha
.lha	application/octet-stream
.lha	application/x-lha
.lhx	application/octet-stream
.list	text/plain
.lma	audio/nspaudio
.lma	audio/x-nspaudio
.log	text/plain
.lsp	application/x-lisp
.lsp	text/x-script.lisp
.lst	text/plain
.lsx	text/x-la-asf
.ltx	application/x-latex
.lzh	application/octet-stream
.lzh	application/x-lzh
.lzx	application/lzx
.lzx	application/octet-stream
.lzx	application/x-lzx
.m	text/plain
.m	text/x-m
.m1v	video/mpeg
.m2a	audio/mpeg
.m2v	video/mpeg
.m3u	audio/x-mpequrl
.man	application/x-troff-man
.map	application/x-navimap
.mar	text/plain
.mbd	application/mbedlet
.mc$	application/x-magic-cap-package-1.0
.mcd	application/mcad
.mcd	application/x-mathcad
.mcf	image/vasa
.mcf	text/mcf
.mcp	application/netmc
.me	application/x-troff-me
.mht	message/rfc822
.mhtml	message/rfc822
.mid	application/x-midi
.mid	audio/midi
.mid	audio/x-mid
.mid	audio/x-midi
.mid	music/crescendo
.mid	x-music/x-midi
.midi	application/x-midi
.midi	audio/midi
.midi	audio/x-mid
.midi	audio/x-midi
.midi	music/crescendo
.midi	x-music/x-midi
.mif	application/x-frame
.mif	application/x-mif
.mime	message/rfc822
.mime	www/mime
.mjf	audio/x-vnd.audioexplosion.mjuicemediafile
.mjpg	video/x-motion-jpeg
.mka	audio/x-matroska
.mkv	video/x-matroska
.mm	application/base64
.mm	application/x-meme
.mme	application/base64
.mod	audio/mod
.mod	audio/x-mod
.moov	video/quicktime
.mov	video/quicktime
.movie	video/x-sgi-movie
.mp2	audio/mpeg
.mp2	audio/x-mpeg
.mp2	video/mpeg
.mp2	video/x-mpeg
.mp2	video/x-mpeq2a
.mp3	audio/mpeg3
.mp3	audio/x-mpeg-3
.mp3	video/mpeg
.mp3	video/x-mpeg
.mp4	video/mp4
.mpa	audio/mpeg
.mpa	video/mpeg
.mpc	application/x-project
.mpe	video/mpeg
.mpeg	video/mpeg
.mpg	audio/mpeg
.mpg	video/mpeg
.mpga	audio/mpeg
.mpp	application/vnd.ms-project
.mpt	application/x-project
.mpv	application/x-project
.mpx	application/x-project
.mrc	application/marc
.ms	application/x-troff-ms
.mv	video/x-sgi-movie
.my	audio/make
.mzz	application/x-vnd.audioexplosion.mzz
.nap	image/naplps
.naplps	image/naplps
.nc	application/x-netcdf
.ncm	application/vnd.nokia.configuration-message
.nif	image/x-niff
.niff	image/x-niff
.nix	application/x-mix-transfer
.nsc	application/x-conference
.nvd	application/x-navidoc
.o	application/octet-stream
.oda	application/oda
.ogg	audio/ogg
.ogg	video/ogg
.omc	application/x-omc
.omcd	application/x-omcdatamaker
.omcr	application/x-omcregerator
.otf	font/otf
.p	text/x-pascal
.p10	application/pkcs10
.p10	application/x-pkcs10
.p12	application/pkcs-12
.p12	application/x-pkcs12
.p7a	application/x-pkcs7-signature
.p7c	application/pkcs7-mime
.p7c	application/x-pkcs7-mime
.p7m	application/pkcs7-mime
.p7m	application/x-pkcs7-mime
.p7r	application/x-pkcs7-certreqresp
.p7s	application/pkcs7-signature
.part	application/pro_eng
.pas	text/pascal
.pbm	image/x-portable-bitmap
.pcl	application/vnd.hp-pcl
.pcl	application/x-pcl
.pct	image/x-pict
.pcx	image/x-pcx
.pdb	chemical/x-pdb
.pdf	application/pdf
.pfunk	audio/make
.pfunk	audio/make.my.funk
.pgm	image/x-portable-graymap
.pgm	image/x-portable-greymap
.pic	image/pict
.pict	image/pict
.pkg	application/x-newton-compatible-pkg
.pko	application/vnd.ms-pki.pko
.pl	text/plain
.pl	text/x-script.perl
.plx	application/x-pixclscript
.pm	image/x-xpixmap
.pm	text/x-script.perl-module
.pm4	application/x-pagemaker
.pm5	application/x-pagemaker
.png	image/png
.pnm	application/x-portable-anymap
.pnm	image/x-portable-anymap
.pot	application/mspowerpoint
.pot	application/vnd.ms-powerpoint
.pov	model/x-pov
.ppa	application/vnd.ms-powerpoint
.ppm	image/x-portable-pixmap
.pps	application/mspowerpoint
.pps	application/vnd.ms-powerpoint
.ppt	application/mspowerpoint
.ppt	application/powerpoint
.ppt	application/vnd.ms-powerpoint
.ppt	application/x-mspowerpoint
.pptx	application/vnd.openxmlformats-officedocument.presentationml.presentation
.ppz	application/mspowerpoint
.pre	application/x-freelance
.prt	application/pro_eng
.ps	application/postscript
.psd	application/octet-stream
.pvu	paleovu/x-pv
.pwz	application/vnd.ms-powerpoint
.py	text/x-script.phyton
.pyc	application/x-bytecode.python
.qcp	audio/vnd.qcelp
.qd3	x-world/x-3dmf
.qd3d	x-world/x-3dmf
.qif	image/x-quicktime
.qt	video/quicktime
.qtc	video/x-qtc
.qti	image/x-quicktime
.qtif	image/x-quicktime
.ra	audio/x-pn-realaudio
.ra	audio/x-pn-realaudio-plugin
.ra	audio/x-realaudio
.ram	audio/x-pn-realaudio
.ras	application/x-cmu-raster
.ras	image/cmu-raster
.ras	image/x-cmu-raster
.rast	image/cmu-raster
.rar	application/vnd.rar
.rexx	text/x-script.rexx
.rf	image/vnd.rn-realflash
.rgb	image/x-rgb
.rm	application/vnd.rn-realmedia
.rm	audio/x-pn-realaudio
.rmi	audio/mid
.rmm	audio/x-pn-realaudio
.rmp	audio/x-pn-realaudio
.rmp	audio/x-pn-realaudio-plugin
.rng	application/ringing-tones
.rng	application/vnd.nokia.ringing-tone
.rnx	application/vnd.rn-realplayer
.roff	application/x-troff
.rp	image/vnd.rn-realpix
.rpm	audio/x-pn-realaudio-plugin
.rt	text/richtext
.rt	text/vnd.rn-realtext
.rtf	application/rtf
.rtf	application/x-rtf
.rtf	text/richtext
.rtx	application/rtf
.rtx	text/richtext
.rv	video/vnd.rn-realvideo
.s	text/x-asm
.s3m	audio/s3m
.saveme	application/octet-stream
.sbk	application/x-tbook
.scm	application/x-lotusscreencam
.scm	text/x-script.guile
.scm	text/x-script.scheme
.scm	video/x-scm
.sdml	text/plain
.sdp	application/sdp
.sdp	application/x-sdp
.sdr	application/sounder
.sea	application/sea
.sea	application/x-sea
.set	application/set
.sgm	text/sgml
.sgm	text/x-sgml
.sgml	text/sgml
.sgml	text/x-sgml
.sh	application/x-bsh
.sh	application/x-sh
.sh	application/x-shar
.sh	text/x-script.sh
.shar	application/x-bsh
.shar	application/x-shar
.shtml	text/html
.shtml	text/x-server-parsed-html
.sid	audio/x-psid
.sit	application/x-sit
.sit	application/x-stuffit
.skd	application/x-koan
.skm	application/x-koan
.skp	application/x-koan
.skt	application/x-koan
.sl	application/x-seelogo
.smi	application/smil
.smil	application/smil
.snd	audio/basic
.snd	audio/x-adpcm
.sol	application/solids
.spc	application/x-pkcs7-certificates
.spc	text/x-speech
.spl	application/futuresplash
.spr	application/x-sprite
.sprite	application/x-sprite
.src	application/x-wais-source
.ssi	text/x-server-parsed-html
.ssm	application/streamingmedia
.sst	application/vnd.ms-pki.certstore
.step	application/step
.stl	application/sla
.stl	application/vnd.ms-pki.stl
.stl	application/x-navistyle
.stp	application/step
.sv4cpio	application/x-sv4cpio
.sv4crc	application/x-sv4crc
.svf	image/vnd.dwg
.svf	image/x-dwg
.svg	image/svg+xml
.svr	application/x-world
.svr	x-world/x-svr
.swf	application/x-shockwave-flash
.t	application/x-troff
.talk	text/x-speech
.tar	application/x-tar
.tbk	application/toolbook
.tbk	application/x-tbook
.tcl	application/x-tcl
.tcl	text/x-script.tcl
.tcsh	text/x-script.tcsh
.tex	application/x-tex
.texi	application/x-texinfo
.texinfo	application/x-texinfo
.text	application/plain
.text	text/plain
.tgz	application/gnutar
.tgz	application/x-compressed
.tif	image/tiff
.tif	image/x-tiff
.tiff	image/tiff
.tiff	image/x-tiff
.tr	application/x-troff
.ts	video/mp2t
.tsi	audio/tsp-audio
.tsp	application/dsptype
.tsp	audio/tsplayer
.tsv	text/tab-separated-values
.turbot	image/florian
.txt	text/plain
.uil	text/x-uil
.uni	text/uri-list
.unis	text/uri-list
.unv	application/i-deas
.uri	text/uri-list
.uris	text/uri-list
.ustar	application/x-ustar
.ustar	multipart/x-ustar
.uu	application/octet-stream
.uu	text/x-uuencode
.uue	text/x-uuencode
.vcd	application/x-cdlink
.vcs	text/x-vcalendar
.vda	application/vda
.vdo	video/vdo
.vew	application/groupwise
.viv	video/vivo
.viv	video/vnd.vivo
.vivo	video/vivo
.vivo	video/vnd.vivo
.vmd	application/vocaltec-media-desc
.vmf	application/vocaltec-media-file
.voc	audio/voc
.voc	audio/x-voc
.vos	video/vosaic
.vox	audio/voxware
.vqe	audio/x-twinvq-plugin
.vqf	audio/x-twinvq
.vql	audio/x-twinvq-plugin
.vrml	application/x-vrml
.vrml	model/vrml
.vrml	x-world/x-vrml
.vrt	x-world/x-vrt
.vsd	application/x-visio
.vst	application/x-visio
.vsw	application/x-visio
.w60	application/wordperfect6.0
.w61	application/wordperfect6.1
.w6w	application/msword
.wav	audio/wav
.wav	audio/x-wav
.wb1	application/x-qpro
.wbmp	image/vnd.wap.wbmp
.web	application/vnd.xara
.webm	video/webm
.webp	image/webp
.wiz	application/msword
.wk1	application/x-123
.wmf	windows/metafile
.wml	text/vnd.wap.wml
.wmlc	application/vnd.wap.wmlc
.wmls	text/vnd.wap.wmlscript
.wmlsc	application/vnd.wap.wmlscriptc
.word	application/msword
.woff	font/woff
.woff2	font/woff2
.wp	application/wordperfect
.wp5	application/wordperfect
.wp5	application/wordperfect6.0
.wp6	application/wordperfect
.wpd	application/wordperfect
.wpd	application/x-wpwin
.wq1	application/x-lotus
.wri	application/mswrite
.wri	application/x-wri
.wrl	application/x-world
.wrl	model/vrml
.wrl	x-world/x-vrml
.wrz	model/vrml
.wrz	x-world/x-vrml
.wsc	text/scriplet
.wsrc	application/x-wais-source
.wtk	application/x-wintalk
.xbm	image/x-xbitmap
.xbm	image/x-xbm
.xbm	image/xbm
.xdr	video/x-amt-demorun
.xgz	xgl/drawing
.xif	image/vnd.xiff
.xl     application/excel
.xla	application/excel
.xla	application/x-excel
.xla	application/x-msexcel
.xlb	application/excel
.xlb	application/vnd.ms-excel
.xlb	application/x-excel
.xlc	application/excel
.xlc	application/vnd.ms-excel
.xlc	application/x-excel
.xld	application/excel
.xld	application/x-excel
.xlk	application/excel
.xlk	application/x-excel
.xll	application/excel
.xll	application/vnd.ms-excel
.xll	application/x-excel
.xlm	application/excel
.xlm	application/vnd.ms-excel
.xlm	application/x-excel
.xls	application/excel
.xls	application/vnd.ms-excel
.xls	application/x-excel
.xls	application/x-msexcel
.xlt	application/excel
.xlt	application/x-excel
.xlv	application/excel
.xlv	application/x-excel
.xlw	application/excel
.xlw	application/vnd.ms-excel
.xlw	application/x-excel
.xlw	application/x-msexcel
.xm	audio/xm
.xml	application/xml
.xml	text/xml
.xmz	xgl/movie
.xpix	application/x-vnd.ls-xpix
.xpm	image/x-xpixmap
.xpm	image/xpm
.x-png	image/png
.xlsx	application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
.xsr	video/x-amt-showrun
.xwd	image/x-xwd
.xwd	image/x-xwindowdump
.xyz	chemical/x-pdb
.yaml	application/x-yaml
.yml	application/x-yaml
.z	application/x-compress
.z	application/x-compressed
.zip	application/x-compressed
.zip	application/x-zip-compressed
.zip	application/zip
.zip	multipart/x-zip
.zoo	application/octet-stream
.zsh	text/x-script.zsh
"""


@st.cache_data
def get_cpcc_css():
    # Embed custom fonts using HTML and CSS
    css = """
        <style>
            @font-face {
                font-family: "Franklin Gothic";
                src: url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.eot");
                src: url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.eot?#iefix")format("embedded-opentype"),
                url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.woff2")format("woff2"),
                url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.woff")format("woff"),
                url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.ttf")format("truetype"),
                url("https://db.onlinewebfonts.com/t/9c9dbb999dd7068f51335d93cc7328bd.svg#Franklin Gothic")format("svg");
            }

            @font-face {
                font-family: 'ITC New Baskerville';
                src: url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.eot");
                src: url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.eot?#iefix")format("embedded-opentype"),
                url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.woff2")format("woff2"),
                url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.woff")format("woff"),
                url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.ttf")format("truetype"),
                url("https://db.onlinewebfonts.com/t/501ade6e29baa5c62c15ec28f3ed2c62.svg#ITC New Baskerville")format("svg");
            }

            body {
                font-family: 'Franklin Gothic', sans-serif;
            }

            h1, h2, h3, h4, h5, h6 {
                font-family: 'Franklin Gothic', sans-serif;
                font-weight: normal;
            }

            p {
                font-family: 'ITC New Baskerville', sans-serif;
                font-weight: normal;
            }
        </style>
        """
    return css


@st.cache_resource(hash_funcs={ChatOpenAI: id})
def get_custom_llm(temperature: float, model: str, service_tier: str = "default") -> ChatOpenAI:
    """
    This function returns a cached instance of ChatOpenAI based on the temperature and model.
    If the temperature or model changes, a new instance will be created and cached.
    """
    return ChatOpenAI(temperature=temperature,
                      model=model,
                      openai_api_key=st.session_state.openai_api_key,
                      use_responses_api=True,
                      service_tier=service_tier
                      # streaming=True
                      )


def get_file_extension_from_filepath(file_path: str, remove_leading_dot: bool = False) -> str:
    basename = os.path.basename(file_path)
    file_name, file_extension = os.path.splitext(basename)
    if remove_leading_dot and file_extension.startswith("."):
        # st.info("Removing leading dot from file extension: " + file_extension)
        file_extension = file_extension[1:]

    if file_extension:
        file_extension = file_extension.lower()

    # st.info("Base Name: " + basename + " | File Name: " + file_name + " | File Extension: " + file_extension)

    return file_extension


def get_unique_extensions() -> list[str]:
    """
    Returns a sorted list of unique file extensions.
    """
    return sorted({ext.lower() for ext in EXTENSION_TO_LANGUAGES.keys()})


def get_unique_languages() -> list[str]:
    """
    Returns a sorted list of unique language names across all extensions.
    """
    return sorted({
        lang.lower()
        for langs in EXTENSION_TO_LANGUAGES.values()
        for lang in langs
    })


def define_code_language_selection(unique_key: str | int, default_option: str = 'java'):
    # Get unique code languages
    code_languages = get_unique_languages()

    # List of available languages
    selected_language = st.selectbox(label="Select Code Language",
                                     key="language_select_" + unique_key,
                                     options=code_languages,
                                     index=code_languages.index(default_option))
    return selected_language


# streamlit model/tier selector with updated GPT-5 Flex pricing (per 1M tokens)
import json
from typing import Dict

import streamlit as st


def define_chatGPTModel(unique_key: str | int,
                        default_min_value: float = 0.2,
                        default_max_value: float = 0.8,
                        default_temp_value: float = 0.2,
                        default_step: float = 0.1,
                        default_option: str = "gpt-5") -> Dict[str, Any]:
    """
    Presents model selection, temperature slider, and service tier.
    Returns JSON-serializable dict:
      {
        "model": str,
        "temperature": float,
        "service_tier": "Standard" | "Priority" | "Flex",
        "pricing": {input, cached, output, unit},
        "token_limits": {"context_window": int, "max_input": int|None, "max_output": int|None}
      }

    Notes
    - Units are per **1M tokens** (matches OpenAI pricing pages).
    - Models listed support structured outputs (JSON/JSON Schema via Responses API).
    """

    uk = str(unique_key)

    # === Models supporting structured output ===
    model_options = [
        # GPT-5 family only (standardized)
        "gpt-5",
        "gpt-5-mini",
        "gpt-5-nano",
    ]
    if default_option not in model_options:
        default_option = "gpt-5"

    # === Indicative context windows ===
    token_limits = {
        # GPT-5 family (API page lists 400K total; 128K max output cap)
        "gpt-5": {"context_window": 400_000, "max_input": 272_000, "max_output": 128_000},
        "gpt-5-mini": {"context_window": 400_000, "max_input": 272_000, "max_output": 128_000},
        "gpt-5-nano": {"context_window": 400_000, "max_input": 272_000, "max_output": 128_000},
    }

    # === Standard pricing (per 1M tokens) ===
    # Source: OpenAI API Pricing page
    standard_prices = {
        "gpt-5": {"input": 1.25, "cached": 0.125, "output": 10.00, "unit": "1M tokens"},
        "gpt-5-mini": {"input": 0.25, "cached": 0.025, "output": 2.00, "unit": "1M tokens"},
        "gpt-5-nano": {"input": 0.05, "cached": 0.005, "output": 0.40, "unit": "1M tokens"},
    }

    # === Priority pricing (per 1M tokens) ===
    # Source: OpenAI "Priority Processing for API Customers"
    priority_prices = {
        "gpt-5": {"input": 2.50, "cached": 0.250, "output": 20.00, "unit": "1M tokens"},
        "gpt-5-mini": {"input": 0.45, "cached": 0.045, "output": 3.60, "unit": "1M tokens"},
    }

    # === Flex pricing (per 1M tokens) — updated from your screenshot ===
    # If you want to override via env, set FLEX_PRICE_OVERRIDES as JSON.
    flex_prices_default = {
        "gpt-5": {"input": 0.625, "cached": 0.0625, "output": 5.00, "unit": "1M tokens"},
        "gpt-5-mini": {"input": 0.125, "cached": 0.0125, "output": 1.00, "unit": "1M tokens"},
        "gpt-5-nano": {"input": 0.025, "cached": 0.0025, "output": 0.20, "unit": "1M tokens"},
        # Leaving 4.x Flex out unless you explicitly want them; easy to add later.
    }
    flex_overrides_env = os.getenv("FLEX_PRICE_OVERRIDES")
    if flex_overrides_env:
        try:
            parsed = json.loads(flex_overrides_env)
            for k, v in parsed.items():
                if isinstance(v, dict):
                    flex_prices_default[k] = {**v, "unit": "1M tokens"}
        except Exception:
            pass

    def get_flex_price(model: str) -> Dict[str, Optional[float]]:
        if model in flex_prices_default:
            d = flex_prices_default[model]
            return {
                "input": d.get("input"),
                "cached": d.get("cached"),
                "output": d.get("output"),
                "unit": d.get("unit", "1M tokens"),
            }
        return {"input": None, "cached": None, "output": None, "unit": "1M tokens"}

    # === UI controls ===
    selected_model = st.selectbox(
        label="Select Model (structured output capable)",
        key=f"chat_select_{uk}",
        options=model_options,
        index=model_options.index(default_option)
    )

    service_tier = st.radio(
        label="Service Tier",
        key=f"chat_tier_{uk}",
        options=["Standard", "Priority", "Flex"],
        index=0
    )

    temperature = st.slider(
        label="Temperature",
        key=f"chat_temp_{uk}",
        min_value=max(default_min_value, 0.0),
        max_value=min(default_max_value, 1.0),
        step=default_step,
        value=default_temp_value,
        format="%.2f"
    )
    if temperature <= 0.3:
        st.caption("Low: most deterministic, best for strict JSON/schema.")
    elif temperature <= 0.7:
        st.caption("Medium: balanced creativity vs. schema adherence.")
    else:
        st.caption("High: diverse outputs; may reduce schema adherence.")

    # === Resolve pricing based on tier ===
    if service_tier == "Standard":
        pricing = standard_prices.get(selected_model,
                                      {"input": None, "cached": None, "output": None, "unit": "1M tokens"})
    elif service_tier == "Priority":
        pricing = priority_prices.get(selected_model,
                                      {"input": None, "cached": None, "output": None, "unit": "1M tokens"})
    else:  # Flex
        pricing = get_flex_price(selected_model)

    # === Display price + limits ===
    tl = token_limits.get(selected_model, {})
    cw = tl.get("context_window")
    max_in = tl.get("max_input")
    max_out = tl.get("max_output")

    def _fmt_price(p: Optional[float], label: str) -> Optional[str]:
        return f"{label}: ${p:.4f} / {pricing['unit']}" if isinstance(p, (int, float)) else None

    parts = [
        _fmt_price(pricing.get("input"), "Input"),
        _fmt_price(pricing.get("cached"), "Cached input"),
        _fmt_price(pricing.get("output"), "Output"),
    ]
    price_line = " | ".join([p for p in parts if p]) if any(parts) else "Pricing: not available"

    cw_bits = [f"Context window: ~{cw:,} tokens" if cw else "Context window: see model docs"]
    if max_in:
        cw_bits.append(f"Max input: ~{max_in:,}")
    if max_out:
        cw_bits.append(f"Max output: ~{max_out:,}")
    st.info(f"Model: {selected_model} | Tier: {service_tier} | {price_line} | {' | '.join(cw_bits)}")

    # === Optional: inline cost estimator ===
    with st.expander("Estimate cost for this request (optional)"):
        in_tokens = st.number_input("Estimated input tokens (prompt)", min_value=0, value=0, step=1000,
                                    key=f"in_tokens_{uk}")
        cached_ratio = st.slider("Estimated % of input tokens served from prompt cache", 0, 100, 0, 5,
                                 key=f"cached_ratio_{uk}")
        out_tokens = st.number_input("Estimated output tokens (completion)", min_value=0, value=0, step=1000,
                                     key=f"out_tokens_{uk}")

        def estimate_cost(pr: Dict[str, Any], in_tok: int, cached_pct: int, out_tok: int) -> Optional[float]:
            if pr.get("input") is None or pr.get("output") is None:
                return None
            cached = pr.get("cached")
            cached_tokens = int(in_tok * (cached_pct / 100.0))
            regular_tokens = max(0, in_tok - cached_tokens)
            per_m = 1_000_000.0
            input_cost = (regular_tokens / per_m) * pr["input"]
            cached_cost = (cached_tokens / per_m) * (cached if cached is not None else pr["input"])
            output_cost = (out_tok / per_m) * pr["output"]
            return round(input_cost + cached_cost + output_cost, 6)

        est = estimate_cost(pricing, in_tokens, cached_ratio, out_tokens)
        if est is None:
            st.warning("Pricing not available for this tier/model combination.")
        else:
            st.success(f"Estimated cost: ${est:,.6f}")

    # Map UI service tiers to LangChain BaseChatOpenAI service_tier values
    _service_tier_map = {"Standard": "default", "Priority": "auto", "Flex": "flex"}

    # ... inside define_chatGPTModel, after service_tier is set:
    langchain_service_tier = _service_tier_map.get(service_tier, "default")

    return {
        "model": selected_model,
        "temperature": float(temperature),
        "service_tier": service_tier,  # UI-facing
        "langchain_service_tier": langchain_service_tier,  # use this when creating ChatOpenAI(...)
        "pricing": pricing,
        "token_limits": {"context_window": cw, "max_input": max_in, "max_output": max_out},
    }


@st.cache_data(ttl=3600)
def _fetch_openrouter_models_cached() -> list:
    """
    Fetch OpenRouter models with Streamlit caching.
    Uses @st.cache_data to cache results for 1 hour.

    This helper runs outside of Streamlit's event loop context,
    allowing asyncio.run() to work properly.

    Returns:
        List of model dictionaries from OpenRouter API
    """
    try:
        import asyncio

        from cqc_cpcc.utilities.AI.openrouter_client import fetch_openrouter_models

        # st.cache_data runs in a context where asyncio.run() is safe
        models = asyncio.run(fetch_openrouter_models())
        logger.info(f"Fetched {len(models)} models from OpenRouter")
        return models
    except Exception as e:
        logger.error(f"Failed to fetch OpenRouter models: {e}", exc_info=True)
        return []


def define_openrouter_model(unique_key: str | int, default_use_auto_route: bool = True) -> Dict[str, Any]:
    """
    Presents OpenRouter model configuration with auto-routing option.
    Returns JSON-serializable dict:
      {
        "use_auto_route": bool,
        "model": str,  # "openrouter/auto" or specific model ID
        "use_openrouter": True,
      }
    
    Args:
        unique_key: Unique key for widget state management
        default_use_auto_route: Default state of auto-routing checkbox
    
    Returns:
        Configuration dictionary for OpenRouter
    """
    uk = str(unique_key)

    # Checkbox for auto-routing (default: True)
    use_auto_route = st.checkbox(
        label="Use Auto Router (Recommended)",
        value=default_use_auto_route,
        key=f"openrouter_auto_{uk}",
        help="Let OpenRouter automatically select the best model for your request"
    )

    selected_model = "openrouter/auto"

    if not use_auto_route:
        # Fetch available models from OpenRouter
        # Using cached helper to avoid asyncio.run() in running event loop
        with st.spinner("Fetching available models from OpenRouter..."):
            models = _fetch_openrouter_models_cached()

            if not models:
                # Fall back to allowed models from environment or default list
                from cqc_cpcc.utilities.env_constants import OPENROUTER_ALLOWED_MODELS

                if OPENROUTER_ALLOWED_MODELS:
                    # Parse comma-separated list from environment
                    allowed_model_ids = [m.strip() for m in OPENROUTER_ALLOWED_MODELS.split(',') if m.strip()]
                    st.info(
                        f"Using allowed models from OPENROUTER_ALLOWED_MODELS environment variable ({len(allowed_model_ids)} models)")
                else:
                    # Use default list of known GPT-5 models
                    allowed_model_ids = [
                        "openai/gpt-5-mini",
                        "openai/gpt-5",
                        "openai/gpt-5-nano"
                    ]
                    st.info("Using default allowed models: GPT-5 family")

                # Create model options from allowed list
                model_options = allowed_model_ids
                model_id_map = {model_id: model_id for model_id in allowed_model_ids}

                # Sort alphabetically
                model_options.sort()

                selected_display = st.selectbox(
                    label="Select OpenRouter Model",
                    key=f"openrouter_model_{uk}",
                    options=model_options,
                    help="Choose a specific model from the allowed models list"
                )

                selected_model = model_id_map.get(selected_display, allowed_model_ids[0])
            else:
                # Create model options from fetched models
                # Format: "model_id - Model Name"
                model_options = []
                model_id_map = {}

                for model in models:
                    model_id = model.get("id", "")
                    model_name = model.get("name", model_id)
                    display_name = f"{model_id} - {model_name}"
                    model_options.append(display_name)
                    model_id_map[display_name] = model_id

                # Sort alphabetically
                model_options.sort()

                selected_display = st.selectbox(
                    label="Select OpenRouter Model",
                    key=f"openrouter_model_{uk}",
                    options=model_options,
                    help="Choose a specific model from OpenRouter's available models"
                )

                selected_model = model_id_map.get(selected_display, "openrouter/auto")

                # Display model information if available
                selected_model_info = next(
                    (m for m in models if m.get("id") == selected_model),
                    None
                )
                if selected_model_info:
                    context_length = selected_model_info.get("context_length", "N/A")
                    pricing = selected_model_info.get("pricing", {})
                    prompt_price = pricing.get("prompt", "N/A")
                    completion_price = pricing.get("completion", "N/A")

                    st.info(
                        f"**Model:** {selected_model}  \n"
                        f"**Context Length:** {context_length:,} tokens  \n"
                        f"**Pricing:** Prompt: ${prompt_price} / token, Completion: ${completion_price} / token"
                    )
    else:
        st.info("**Auto Router:** OpenRouter will automatically select the best model for your request.")

    return {
        "use_auto_route": use_auto_route,
        "model": selected_model,
        "use_openrouter": True,
    }


def reset_session_key_value(key: str):
    st.session_state[key] = str(randint(1000, 100000000))


# Type alias for return type
UploadedFileResult = Union[list[tuple[Any, str]], tuple[Any, str], tuple[None, None]]


def add_upload_file_element(
        uploader_text: str,
        accepted_file_types: list[str],
        success_message: bool = True,
        accept_multiple_files: bool = False,
        key_prefix: str = ""
) -> UploadedFileResult:
    """Add a file uploader element with unique key generation.
    
    Args:
        uploader_text: Label for the file uploader
        accepted_file_types: List of accepted file extensions
        success_message: Whether to show success message on upload
        accept_multiple_files: Whether to accept multiple files
        key_prefix: Prefix for widget keys to ensure uniqueness across contexts
        
    Returns:
        If accept_multiple_files=True: List of (original_name, temp_path) tuples
        If accept_multiple_files=False: Single (original_name, temp_path) tuple
        If no files uploaded: (None, None)
    """
    # Button to reset the multi file uploader
    reset_label = "Reset " + uploader_text + " File Uploader"
    reset_key = key_prefix + reset_label.replace(" ", "_")

    if reset_key not in st.session_state:
        reset_session_key_value(reset_key)

    # Create compound widget key using both context-specific prefix and random value
    # This ensures global uniqueness even if random numbers collide across contexts
    widget_key = f"{reset_key}_{st.session_state[reset_key]}"

    uploaded_files = st.file_uploader(label=uploader_text, type=accepted_file_types,
                                      accept_multiple_files=accept_multiple_files, key=widget_key)

    if accept_multiple_files:
        if st.button("Remove All Files", key="Checkbox_" + widget_key):
            reset_session_key_value(reset_key)
            st.rerun()

        uploaded_file_paths = []
        for uploaded_file in uploaded_files:
            if uploaded_file is not None:
                # Get the original file name
                original_file_name = uploaded_file.name

                # Create a temporary file to store the uploaded file
                temp_file_name = upload_file_to_temp_path(uploaded_file)

                uploaded_file_paths.append((original_file_name, temp_file_name))
        if uploaded_files and success_message:
            st.success("File(s) uploaded successfully.")
        return uploaded_file_paths

    elif uploaded_files is not None:
        # Get the original file name
        original_file_name = uploaded_files.name
        # Create a temporary file to store the uploaded file
        temp_file_name = upload_file_to_temp_path(uploaded_files)

        if success_message:
            st.success("File uploaded successfully.")
        return original_file_name, temp_file_name
    else:
        return None, None


def add_flexible_upload_element(
        uploader_text: str,
        accepted_file_types: list[str],
        success_message: bool = True,
        accept_multiple_files: bool = False,
        key_prefix: str = "",
        allow_url: bool = True,
        allow_brightspace: bool = False
) -> UploadedFileResult:
    """
    Add a flexible file input element supporting local uploads, Google Drive URLs, and direct URLs.

    This function extends add_upload_file_element with URL download capabilities.

    Args:
        uploader_text: Label for the file input
        accepted_file_types: List of accepted file extensions
        success_message: Whether to show success message on upload
        accept_multiple_files: Whether to accept multiple files
        key_prefix: Prefix for widget keys to ensure uniqueness
        allow_url: Whether to show URL input option
        allow_brightspace: Whether to show the "From BrightSpace URL" tab (made the
            first/default tab when enabled). Builds a submissions ZIP from a
            BrightSpace Assignment or Quiz URL with a preview/edit step.

    Returns:
        If accept_multiple_files=True: List of (original_name, temp_path) tuples
        If accept_multiple_files=False: Single (original_name, temp_path) tuple
        If no files provided: (None, None)
    """
    # Create tabs for different input methods. When BrightSpace is enabled it
    # becomes the first/default tab.
    tab_bs = None
    if allow_brightspace and allow_url:
        tab_bs, tab1, tab2 = st.tabs(
            ["🎓 From BrightSpace URL", "📁 Upload File(s)", "🔗 From URL/Google Drive"])
    elif allow_url:
        tab1, tab2 = st.tabs(["📁 Upload File(s)", "🔗 From URL/Google Drive"])
    else:
        tab1 = st.container()
        tab2 = None

    # BrightSpace tab takes precedence when it has produced a confirmed ZIP.
    if tab_bs is not None:
        with tab_bs:
            bs_result = add_brightspace_submission_element(
                accepted_file_types=accepted_file_types,
                key_prefix=key_prefix + "bs_",
                success_message=success_message,
            )
        if bs_result and bs_result.get("path"):
            file_tuple = (bs_result["name"], bs_result["path"])
            return [file_tuple] if accept_multiple_files else file_tuple

    with tab1:
        # Use existing upload function for file uploads
        result = add_upload_file_element(
            uploader_text=uploader_text,
            accepted_file_types=accepted_file_types,
            success_message=success_message,
            accept_multiple_files=accept_multiple_files,
            key_prefix=key_prefix + "local_"
        )

        # If we got files from upload, return them
        if accept_multiple_files:
            if result and len(result) > 0:
                return result
        else:
            if result[0] is not None:
                return result

    if allow_url and tab2:
        with tab2:
            st.markdown("**Paste a URL to download:**")
            st.markdown("- Google Drive share link (e.g., `https://drive.google.com/file/d/FILE_ID/view`)")
            st.markdown("- Direct download URL (e.g., `https://example.com/file.pdf`)")

            # URL input
            url_key = f"{key_prefix}url_input_{uploader_text.replace(' ', '_')}"
            url = st.text_input(
                "File URL",
                key=url_key,
                placeholder="https://drive.google.com/file/d/FILE_ID/view",
                help="Enter a Google Drive share link or direct download URL"
            )

            if url:
                download_button_key = f"{key_prefix}download_btn_{uploader_text.replace(' ', '_')}"
                if st.button("📥 Download from URL", key=download_button_key):
                    with st.spinner("Downloading file..."):
                        result = download_file_from_url(url)

                        if result:
                            original_name, temp_path = result
                            if success_message:
                                st.success(f"✅ Downloaded: {original_name}")

                            # Store in session state for persistence
                            state_key = f"{key_prefix}downloaded_file_{uploader_text.replace(' ', '_')}"
                            st.session_state[state_key] = result

                            if accept_multiple_files:
                                return [result]
                            else:
                                return result

            # Check if we have a previously downloaded file in session state
            state_key = f"{key_prefix}downloaded_file_{uploader_text.replace(' ', '_')}"
            if state_key in st.session_state and st.session_state[state_key]:
                original_name, temp_path = st.session_state[state_key]

                # Show current downloaded file
                st.info(f"📄 Current file: {original_name}")

                # Add clear button
                clear_key = f"{key_prefix}clear_btn_{uploader_text.replace(' ', '_')}"
                if st.button("🗑️ Clear", key=clear_key):
                    del st.session_state[state_key]
                    st.rerun()

                if accept_multiple_files:
                    return [st.session_state[state_key]]
                else:
                    return st.session_state[state_key]

    # No files from either source
    if accept_multiple_files:
        return []
    else:
        return None, None


class _BrightSpaceJob:
    """Runs the BrightSpace fetch on a background thread (Selenium is blocking).

    The thread must not touch Streamlit APIs / session_state. It writes only to
    plain attributes here and to the thread-safe ``MfaBridge``; the Streamlit
    script polls those on rerun.
    """

    def __init__(self, url: str, accepted_file_types: list[str], bridge):
        import threading
        self.url = url
        self.accepted_file_types = accepted_file_types
        self.bridge = bridge
        self.result = None  # BrightSpaceFetchResult
        self.error: Optional[str] = None
        self.done = threading.Event()
        self._progress: list[str] = []
        self._lock = threading.Lock()
        self._thread = threading.Thread(target=self._run, daemon=True)

    def start(self):
        self._thread.start()

    def _record(self, message: str):
        with self._lock:
            self._progress.append(message)

    def latest_progress(self) -> Optional[str]:
        with self._lock:
            return self._progress[-1] if self._progress else None

    def _run(self):
        from cqc_cpcc.utilities.brightspace_submissions import (
            build_submissions_zip_from_brightspace_url,
        )
        try:
            self.result = build_submissions_zip_from_brightspace_url(
                self.url,
                self.accepted_file_types,
                progress=self._record,
                mfa_handler=self.bridge,
            )
        except Exception as e:  # noqa: BLE001 - surfaced to the UI
            self.error = str(e)
            logger.exception("BrightSpace fetch failed")
        finally:
            self.done.set()


class _BrightSpaceWritebackJob:
    """Runs the BrightSpace draft grade write-back on a background thread.

    Like ``_BrightSpaceJob``, the thread must not touch Streamlit APIs; it writes
    only to plain attributes + the thread-safe ``MfaBridge``. ``dry_run`` (default
    True) navigates and locates the write targets but never fills or saves.
    """

    def __init__(self, url: str, items: list, bridge, dry_run: bool = True):
        import threading
        self.url = url
        self.items = items
        self.bridge = bridge
        self.dry_run = dry_run
        self.report = None  # GradeWriteReport
        self.error: Optional[str] = None
        self.done = threading.Event()
        self._progress: list[str] = []
        self._lock = threading.Lock()
        self._thread = threading.Thread(target=self._run, daemon=True)

    def start(self):
        self._thread.start()

    def _record(self, message: str):
        with self._lock:
            self._progress.append(message)

    def latest_progress(self) -> Optional[str]:
        with self._lock:
            return self._progress[-1] if self._progress else None

    def _run(self):
        from cqc_cpcc.utilities.brightspace_writeback import push_grades_to_brightspace
        try:
            self.report = push_grades_to_brightspace(
                self.url, self.items, progress=self._record,
                mfa_handler=self.bridge, dry_run=self.dry_run,
            )
        except Exception as e:  # noqa: BLE001 - surfaced to the UI
            self.error = str(e)
            logger.exception("BrightSpace write-back failed")
        finally:
            self.done.set()


def _render_mfa_prompt(bridge) -> None:
    """Show the MFA number-matching prompt prominently (number + screenshot).

    Rendered inline rather than as a modal dialog because the running-job view
    auto-refreshes via a poll/rerun loop, which does not compose with a modal.
    """
    challenge = bridge.challenge
    if challenge is None:
        return
    with st.container(border=True):
        st.subheader("🔐 Two-factor approval needed")
        if challenge.number:
            st.markdown("### Enter this number in your Authenticator app, then approve:")
            # Large, prominent number for easy reading on the same page.
            st.markdown(
                f"<div style='font-size:4rem;font-weight:700;letter-spacing:0.25rem;"
                f"text-align:center;padding:0.5rem 0;'>{challenge.number}</div>",
                unsafe_allow_html=True,
            )
        else:
            st.info(
                "A sign-in approval is pending — reading the matching number from the "
                "login page… Approve the push on your device (the number will appear "
                "here in a moment if your app asks for one)."
            )
        if challenge.message:
            st.caption(challenge.message)
        if challenge.screenshot_png:
            with st.expander("Show live browser view", expanded=not challenge.number):
                st.image(
                    challenge.screenshot_png,
                    caption="Live browser view",
                    use_container_width=True,
                )


def add_brightspace_submission_element(
        accepted_file_types: list[str],
        key_prefix: str = "",
        success_message: bool = True,
) -> Optional[dict]:
    """Render the "From BrightSpace URL" input + preview/edit, returning results.

    A single fetch yields both the student submissions ZIP and the scraped
    instructions (assignment description or first quiz question). Returns a dict
    ``{"name", "path", "instructions", "route"}`` once the user confirms via the
    preview/edit panel, otherwise ``None``.
    """
    import time

    job_key = key_prefix + "job"
    fetch_result_key = key_prefix + "fetch_result"
    confirmed_key = key_prefix + "confirmed"

    # If the user already confirmed a generated ZIP, keep returning it.
    if st.session_state.get(confirmed_key):
        confirmed = st.session_state[confirmed_key]
        st.success(f"✅ Using generated file: {confirmed['name']}")
        if confirmed.get("instructions"):
            st.caption("📝 Instructions captured from this URL will be used below.")

        # Persisted review: let the instructor re-open the file picker and
        # select/de-select student folders/files after confirming, without
        # re-fetching. Operates on the ORIGINAL fetched ZIP (full set) and rebuilds
        # the confirmed ZIP. Route-agnostic (assignment + quiz).
        orig = st.session_state.get(fetch_result_key)
        if orig is not None and getattr(orig, "zip_path", None):
            with st.expander("📂 Review / change selected files", expanded=False):
                keep_arcs = _render_zip_keep_table(orig.zip_path, key_prefix + "review_")
                if st.button("✅ Apply selection", key=key_prefix + "reapply"):
                    if keep_arcs:
                        new_path = _rebuild_zip_from_kept(orig.zip_path, keep_arcs)
                        ok, detail = _validate_submissions_zip(new_path, keep_arcs)
                        if ok:
                            confirmed = {**confirmed, "path": new_path}
                            st.session_state[confirmed_key] = confirmed
                            st.success(f"Updated: {detail} student submission(s).")
                            st.rerun()
                        else:
                            st.error(f"Selection failed validation: {detail}")
                    else:
                        st.error("Select at least one file to keep.")

        if st.button("🔄 Start over", key=key_prefix + "startover"):
            for k in (job_key, fetch_result_key, confirmed_key):
                st.session_state.pop(k, None)
            st.rerun()
        return confirmed

    st.markdown("**Paste a BrightSpace Assignment or Quiz URL:**")
    st.caption(
        "Assignments use BrightSpace's native download; quizzes collect each "
        "student's file-upload answers. Only the last attempt is kept (review below)."
    )
    url = st.text_input(
        "BrightSpace URL",
        key=key_prefix + "url",
        placeholder="https://brightspace.cpcc.edu/d2l/lms/dropbox/...",
    )

    job = st.session_state.get(job_key)
    fetch_clicked = st.button(
        "📥 Fetch submissions", key=key_prefix + "fetch", disabled=not url
    )

    if fetch_clicked and url:
        from cqc_cpcc.utilities.selenium_util import MfaBridge
        bridge = MfaBridge()
        job = _BrightSpaceJob(url, accepted_file_types, bridge)
        job.start()
        st.session_state[job_key] = job
        st.session_state.pop(fetch_result_key, None)
        st.rerun()

    # Job in progress: show status + MFA prompt, then poll.
    if job is not None and not job.done.is_set():
        status_msg = job.latest_progress() or "Starting..."
        st.info(f"⏳ {status_msg}")
        _render_mfa_prompt(job.bridge)
        if st.button("✖ Cancel", key=key_prefix + "cancel"):
            job.bridge.cancel()
            st.session_state.pop(job_key, None)
            st.rerun()
        # Poll: brief sleep then rerun to refresh progress / MFA number.
        time.sleep(1.5)
        st.rerun()

    # Job finished.
    if job is not None and job.done.is_set():
        if job.error:
            st.error(f"❌ Fetch failed: {job.error}")
            if st.button("Try again", key=key_prefix + "retry"):
                st.session_state.pop(job_key, None)
                st.rerun()
            return None
        if fetch_result_key not in st.session_state:
            st.session_state[fetch_result_key] = job.result

    fetch_result = st.session_state.get(fetch_result_key)
    if fetch_result is not None:
        confirmed = render_zip_preview_editor(fetch_result, key_prefix)
        if confirmed:
            st.session_state[confirmed_key] = confirmed
            st.session_state.pop(job_key, None)
            st.rerun()
    return None


def add_brightspace_source_element(
        accepted_file_types: list[str],
        key_prefix: str = "",
        success_message: bool = True,
) -> Optional[dict]:
    """Top-level BrightSpace fetch: one URL yields BOTH submissions + instructions.

    Thin wrapper over ``add_brightspace_submission_element`` intended to be placed
    near the top of a grading flow so the confirmed result can auto-fill both the
    Assignment Instructions and Student Submissions steps. Returns the confirmed
    dict ``{"name", "path", "instructions", "route"}`` once the user clicks
    "Use the generated file", otherwise ``None``.
    """
    return add_brightspace_submission_element(
        accepted_file_types=accepted_file_types,
        key_prefix=key_prefix,
        success_message=success_message,
    )


def _render_writeback_report(report) -> None:
    """Render a GradeWriteReport: per-student outcomes + unmatched lists."""
    saved = report.saved_count
    mode = "DRY RUN — nothing was saved" if report.dry_run else f"saved {saved} draft(s)"
    st.success(f"Write-back ({report.route}) complete — matched "
               f"{report.matched_count}/{len(report.outcomes)} shown · {mode}.")

    if report.outcomes:
        rows = [{
            "Student": o.display_name,
            "Score to write": o.score_written if o.score_written is not None else "—",
            "Fields found": "✅" if o.fields_found else "❌",
            "Saved draft": "✅" if o.saved else ("—" if report.dry_run else "❌"),
            "Note": o.note,
        } for o in report.outcomes]
        st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

    if report.unmatched_students:
        st.warning("⚠️ Graded students with NO matching BrightSpace learner (skipped): "
                   + ", ".join(report.unmatched_students))
    if report.unmatched_learners:
        st.caption("BrightSpace learners with no grading result: "
                   + ", ".join(report.unmatched_learners))
    for w in report.warnings:
        st.caption(f"• {w}")


def add_brightspace_writeback_element(
        results: list,
        key_prefix: str = "wb_",
        default_url: str = "",
) -> None:
    """Render the "Write grades back to BrightSpace (draft)" panel.

    Pushes each graded student's buffered score + composed feedback onto their
    BrightSpace evaluation page and saves it as a DRAFT (never published). Defaults to
    a safe **dry run** that locates the write targets without filling or saving.

    Args:
        results: ``list[tuple[student_id, RubricAssessmentResult]]`` from
            ``st.session_state.grading_results_by_key[run_key]``.
        key_prefix: Session-state key prefix (unique per grading run).
        default_url: Pre-fill the BrightSpace URL (e.g. the fetched source URL).
    """
    import time
    from cqc_cpcc.utilities.brightspace_writeback import (
        build_write_items_from_results, DEFAULT_SCORE_BUFFER_PCT,
    )

    if not results:
        return

    job_key = key_prefix + "job"
    report_key = key_prefix + "report"

    st.markdown("#### 📤 Write grades back to BrightSpace (draft)")
    st.caption(
        "Pushes each student's score + feedback onto their BrightSpace evaluation "
        "page and **saves as a draft** (never published) so you can review, then "
        "publish later. Start with a dry run to confirm matches and field targets."
    )

    col1, col2 = st.columns(2)
    with col1:
        buffer_pct = st.number_input(
            "Error-buffer % added to each score (capped at max)",
            min_value=0.0, max_value=100.0, value=float(DEFAULT_SCORE_BUFFER_PCT),
            step=1.0, key=key_prefix + "buffer",
            help="Adds this percent of the max points to each score before writing. "
                 "0 disables the buffer.",
        )
    with col2:
        include_criteria = st.checkbox(
            "Include per-criterion feedback", value=True,
            key=key_prefix + "include_criteria",
        )

    url = st.text_input(
        "BrightSpace Assignment or Quiz URL (the one these grades belong to)",
        value=default_url, key=key_prefix + "url",
        placeholder="https://brightspace.cpcc.edu/d2l/lms/...",
    )

    job = st.session_state.get(job_key)

    def _launch(dry_run: bool):
        from cqc_cpcc.utilities.selenium_util import MfaBridge
        items = build_write_items_from_results(
            results, buffer_pct=buffer_pct, include_criteria_feedback=include_criteria,
        )
        bridge = MfaBridge()
        new_job = _BrightSpaceWritebackJob(url, items, bridge, dry_run=dry_run)
        new_job.start()
        st.session_state[job_key] = new_job
        st.session_state.pop(report_key, None)
        st.rerun()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔍 Preview write (dry run)", key=key_prefix + "dry",
                     disabled=not url, use_container_width=True):
            _launch(dry_run=True)
    with c2:
        confirm = st.checkbox("I reviewed the dry run — write drafts for real",
                              key=key_prefix + "confirm")
        if st.button("✍️ Write drafts to BrightSpace", key=key_prefix + "real",
                     disabled=not (url and confirm), use_container_width=True):
            _launch(dry_run=False)

    # Job in progress: status + MFA prompt, then poll.
    if job is not None and not job.done.is_set():
        st.info(f"⏳ {job.latest_progress() or 'Starting...'}")
        _render_mfa_prompt(job.bridge)
        if st.button("✖ Cancel", key=key_prefix + "cancel"):
            job.bridge.cancel()
            st.session_state.pop(job_key, None)
            st.rerun()
        time.sleep(1.5)
        st.rerun()

    if job is not None and job.done.is_set():
        if job.error:
            st.error(f"❌ Write-back failed: {job.error}")
            if st.button("Try again", key=key_prefix + "retry"):
                st.session_state.pop(job_key, None)
                st.rerun()
            return
        if report_key not in st.session_state:
            st.session_state[report_key] = job.report
            st.session_state.pop(job_key, None)
            st.rerun()

    report = st.session_state.get(report_key)
    if report is not None:
        _render_writeback_report(report)


def _render_zip_keep_table(zip_path: str, key_prefix: str) -> Optional[set]:
    """Render an editable keep/deselect table for every file in ``zip_path``.

    Returns the set of kept arcnames, or ``None`` if the ZIP holds no files. Shared
    by the initial preview and the post-confirm "review files" panel so the picker is
    identical wherever it appears.
    """
    with zipfile.ZipFile(zip_path, "r") as zf:
        arcnames = [i.filename for i in zf.infolist() if not i.is_dir()]
    if not arcnames:
        st.error("The generated ZIP is empty.")
        return None

    rows = []
    for arc in arcnames:
        folder, _, fname = arc.partition("/")
        rows.append({"keep": True, "student folder": folder, "file": fname, "_arc": arc})
    df = pd.DataFrame(rows)

    edited = st.data_editor(
        df,
        key=key_prefix + "editor",
        column_config={
            "keep": st.column_config.CheckboxColumn("keep", help="Uncheck to remove"),
            "student folder": st.column_config.TextColumn(disabled=True),
            "file": st.column_config.TextColumn(disabled=True),
            "_arc": None,  # hidden
        },
        hide_index=True,
        use_container_width=True,
    )

    kept = edited[edited["keep"]]
    n_folders = kept["student folder"].nunique()
    st.caption(f"Will include {len(kept)} file(s) across {n_folders} student folder(s).")
    return set(kept["_arc"].tolist())


def _rebuild_zip_from_kept(src_zip_path: str, keep_arcs: set) -> str:
    """Write a new ZIP containing only ``keep_arcs`` from ``src_zip_path``.

    Preserves the original folder/file arcnames so the grader's extractor still sees
    one folder per student. Returns the new ZIP's path.
    """
    out = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    out.close()
    with zipfile.ZipFile(src_zip_path, "r") as src, \
            zipfile.ZipFile(out.name, "w", zipfile.ZIP_DEFLATED) as dst:
        for arc in keep_arcs:
            with src.open(arc) as f:
                dst.writestr(arc, f.read())
    return out.name


def _validate_submissions_zip(zip_path: str, keep_arcs: set) -> tuple[bool, str]:
    """Check a rebuilt submissions ZIP parses the way the grader expects.

    Returns ``(True, "<count>")`` with the student-submission count on success, or
    ``(False, "<error message>")`` on failure.
    """
    kept_extensions = sorted({
        os.path.splitext(arc)[1].lstrip(".").lower()
        for arc in keep_arcs if os.path.splitext(arc)[1]
    })
    try:
        from cqc_cpcc.utilities.zip_grading_utils import (
            extract_student_submissions_from_zip,
        )
        submissions = extract_student_submissions_from_zip(zip_path, kept_extensions)
        return True, str(len(submissions))
    except Exception as e:  # noqa: BLE001
        return False, str(e)


def render_zip_preview_editor(fetch_result, key_prefix: str) -> Optional[dict]:
    """Preview a generated submissions ZIP and let the user prune folders/files.

    Also previews the scraped instructions (editable) so one BrightSpace fetch
    confirms both the submissions and the assignment instructions.

    Returns a dict ``{"name", "path", "instructions", "route"}`` when the user
    clicks "Use the generated file" (a new ZIP rebuilt from the kept entries),
    otherwise ``None``.
    """
    zip_path = fetch_result.zip_path

    st.markdown(f"**Preview generated submissions** — route: `{fetch_result.route}`")
    for warning in getattr(fetch_result, "warnings", []) or []:
        st.warning(f"⚠️ {warning}")

    # Instructions captured from the same URL (assignment description or first
    # quiz question). Editable so the instructor can correct/trim before grading.
    scraped_instructions = getattr(fetch_result, "instructions", None)
    instructions_key = key_prefix + "instructions_text"
    if scraped_instructions:
        st.markdown("**Captured instructions** (edit if needed):")
        edited_instructions = st.text_area(
            "Assignment instructions",
            value=st.session_state.get(instructions_key, scraped_instructions),
            key=instructions_key,
            height=160,
        )
    else:
        edited_instructions = st.session_state.get(instructions_key)
        st.info(
            "No instructions were auto-captured from this URL. You can still grade "
            "submissions — add instructions in the Assignment Instructions section."
        )

    # Editable keep/deselect table of every file in the ZIP (default: keep all).
    keep_arcs = _render_zip_keep_table(zip_path, key_prefix)
    if keep_arcs is None:
        return None

    col1, col2 = st.columns(2)
    use_clicked = col1.button("✅ Use the generated file", key=key_prefix + "use")
    refetch_clicked = col2.button("🔄 Re-fetch", key=key_prefix + "refetch")

    if refetch_clicked:
        for k in ("job", "fetch_result", "confirmed"):
            st.session_state.pop(key_prefix + k, None)
        st.rerun()

    if use_clicked:
        if not keep_arcs:
            st.error("Select at least one file to keep.")
            return None

        out_path = _rebuild_zip_from_kept(zip_path, keep_arcs)
        ok, detail = _validate_submissions_zip(out_path, keep_arcs)
        if ok:
            st.success(f"✅ Generated file ready: {detail} student submission(s).")
        else:
            st.error(f"Generated ZIP failed validation: {detail}")
            return None

        return {
            "name": "brightspace_submissions.zip",
            "path": out_path,
            "instructions": (edited_instructions or "").strip() or None,
            "route": getattr(fetch_result, "route", None),
        }

    return None


def upload_file_to_temp_path(uploaded_file: UploadedFile):
    file_extension = get_file_extension_from_filepath(uploaded_file.name)

    # Create a temporary file to store the uploaded instructions
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_extension)
    temp_file.write(uploaded_file.getvalue())
    # temp_file.close()

    return temp_file.name


def process_file(file_path, allowed_file_extensions):
    """ Using a file path determine if the file is a zip or single file and gives the contents back if single or dict mapping the studnet name and timestamp back to the combined contents"""

    # If it's a zip file
    if file_path.endswith('.zip'):
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            folder_contents = {}
            for zip_info in zip_file.infolist():
                if any(zip_info.filename.lower().endswith(ext) for ext in allowed_file_extensions):
                    folder_path = os.path.dirname(zip_info.filename)
                    with zip_file.open(zip_info) as file:
                        file_contents = file.read()
                    folder_contents.setdefault(folder_path, []).append(file_contents)

            for folder_path, files in folder_contents.items():
                concatenated_contents = b''.join(files)
                logger.debug("Contents of folder '%s': %s", folder_path, concatenated_contents.decode())

    # If it's a single file
    else:
        if any(file_path.lower().endswith(ext) for ext in allowed_file_extensions):
            with open(file_path, 'r') as file:
                logger.debug("Contents of single file: %s", file.read())


def choose_preferred_mime(mime_list):
    # Define a priority order for MIME types
    priority_order = [
        "application/octet-stream",
        "application/zip"

    ]

    for mime in priority_order:
        if mime in mime_list:
            return mime

    # Return the first MIME type if none match the priority order
    return mime_list[0]


def get_file_mime_type(file_extension: str):
    # Check if file_extension is prefixed with "." if not add it first
    if not file_extension.startswith("."):
        file_extension = "." + file_extension

    # Define the mapping of file extensions to MIME types
    mime_dict = {}
    lines = mime_types_str.strip().split('\n')
    for line in lines:
        try:
            key, value = line.split()
        except ValueError:
            logger.warning("Error splitting MIME mapping line: %s", line)
            key = None
            value = None

        if key in mime_dict:
            mime_dict[key].append(value)
        else:
            mime_dict[key] = [value]

    # Create a dictionary with preferred MIME types
    preferred_mime_dict = {ext: choose_preferred_mime(mimes) for ext, mimes in mime_dict.items()}

    return preferred_mime_dict.get(file_extension, "application/octet-stream")


def on_download_click(download_button_placeholder: DeltaGenerator, file_path: str, button_label: str,
                      download_file_name: str):
    file_extension = get_file_extension_from_filepath(download_file_name)
    mime_type = get_file_mime_type(file_extension)
    # st.info("file_extension: " + file_extension + " | mime_type: " + mime_type)

    # file_content = read_file(file_path)
    # Read the content of the file
    with open(file_path, "rb") as file:
        file_content = file.read()

    # st.info("file_path: "+file_path+" | download_file_name: "+download_file_name)
    # st.markdown(file_content)

    # Trigger the download of the file
    download_button_placeholder.download_button(label=button_label, data=file_content,
                                                file_name=download_file_name, mime=mime_type
                                                # , key=download_file_name
                                                )


def create_zip_file(file_paths: list[tuple[str, str]]) -> str:
    # Create a temporary file to store the zip file
    zip_file = tempfile.NamedTemporaryFile(delete=False)
    zip_file.close()  # Close the file to use it as the output path for the zip file

    with zipfile.ZipFile(zip_file.name, 'w') as zipf:
        for orig_file_path, temp_file_path in file_paths:
            # Get the base file name from the original file path
            base_file_name = os.path.basename(orig_file_path)
            # Add the temporary file to the zip file with the original file name
            zipf.write(temp_file_path, arcname=base_file_name)

    # Return the path of the zip file
    return zip_file.name




def export_grading_summary_to_excel(
        summary_df: pd.DataFrame,
        include_csv: bool = False
) -> tuple[str, Optional[str]]:
    """
    Export grading summary dataframe to Excel file with professional formatting.
    
    Formats the Excel file with:
    - Header row highlighted with background color
    - Alternating row colors for better readability
    - Auto-fit column widths based on content
    - Proper alignment and borders
    
    Args:
        summary_df: pandas DataFrame with grading summary data
        include_csv: If True, also generate CSV version and return both paths
        
    Returns:
        Tuple of (excel_file_path, csv_file_path) if include_csv=True
        Otherwise tuple of (excel_file_path, None)
    """
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Grading Summary"

    # Define colors for header and alternating rows
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Light blue for alternating rows
    alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Alignment settings
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Add header row
    for col_idx, column_title in enumerate(summary_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Add data rows with alternating colors
    for row_idx, row in enumerate(summary_df.itertuples(index=False), 2):
        # Determine if this is an even or odd row for alternating colors
        if row_idx % 2 == 0:
            row_fill = alt_row_fill
        else:
            row_fill = None

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Apply fill color if alternating row
            if row_fill:
                cell.fill = row_fill

            # Set alignment - center for numbers/percentages, left for text
            if isinstance(value, (int, float)):
                cell.alignment = center_alignment
                # Format percentages if they contain %
            elif isinstance(value, str) and "%" in str(value):
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # Auto-fit column widths
    for col_idx, column_title in enumerate(summary_df.columns, 1):
        max_length = len(str(column_title))

        # Check all values in the column
        for row_idx in range(2, len(summary_df) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save Excel file
    excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    excel_file.close()
    wb.save(excel_file.name)

    csv_file = None
    if include_csv:
        # Save CSV version as well
        csv_file = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
        csv_file.close()
        summary_df.to_csv(csv_file.name, index=False)

    return excel_file.name, csv_file.name


def add_grading_summary_to_zip(
        zip_file_path: str,
        summary_df: pd.DataFrame,
        include_csv: bool = True
) -> str:
    """
    Add grading summary Excel file to an existing zip file.
    
    Creates an Excel file from the summary dataframe with professional formatting,
    then adds it to the zip file. Optionally includes CSV version.
    
    Args:
        zip_file_path: Path to existing zip file
        summary_df: pandas DataFrame with grading summary data
        include_csv: If True, also add CSV version to zip
        
    Returns:
        Path to updated zip file with grading summary included
    """
    # Export summary to Excel (and optionally CSV)
    excel_file_path, csv_file_path = export_grading_summary_to_excel(
        summary_df,
        include_csv=include_csv
    )

    # Create new zip file with all contents
    new_zip_file = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    new_zip_file.close()

    # Copy original zip contents and add summary files
    with zipfile.ZipFile(zip_file_path, 'r') as original_zip:
        with zipfile.ZipFile(new_zip_file.name, 'w') as new_zip:
            # Copy all files from original zip
            for item in original_zip.infolist():
                data = original_zip.read(item.filename)
                new_zip.writestr(item, data)

            # Add Excel summary as primary export
            new_zip.write(excel_file_path, arcname="Grading_Summary.xlsx")

            # Add CSV version if requested
            if include_csv and csv_file_path:
                new_zip.write(csv_file_path, arcname="Grading_Summary.csv")

    # Clean up temporary files
    try:
        os.unlink(excel_file_path)
        if csv_file_path and os.path.exists(csv_file_path):
            os.unlink(csv_file_path)
    except Exception:
        pass

    # Remove the old zip and rename new one
    try:
        os.unlink(zip_file_path)
    except Exception:
        pass

    return new_zip_file.name


def prefix_content_file_name(filename: str, content: str):
    """Prefix content with a filename header.
    
    Adds a comment-style header with the filename at the start
    of the content for better context when displaying code/files.
    
    Args:
        filename: Name of the file to prefix
        content: Content to prefix
        
    Returns:
        Content with filename header prepended
    """
    return f"// File: {filename}\n{content}"


def render_openai_debug_panel(
        correlation_id: str | None = None,
        error: Exception | None = None,
) -> None:
    """Render OpenAI debug panel in Streamlit UI when debug mode is enabled.

    Shows request/response details, correlation ID, and decision notes for
    troubleshooting OpenAI API calls.

    Args:
        correlation_id: Correlation ID for the request (if available)
        error: Exception that occurred (if any)
    """
    import json

    from cqc_cpcc.utilities.AI.openai_debug import get_debug_context
    from cqc_cpcc.utilities.AI.openai_exceptions import (
        OpenAISchemaValidationError,
        OpenAITransportError,
    )
    from cqc_cpcc.utilities.env_constants import CQC_OPENAI_DEBUG

    # Only show debug panel if debug mode is enabled
    if not CQC_OPENAI_DEBUG:
        return

    # Create collapsible debug panel
    with st.expander("🔍 OpenAI Debug Information", expanded=False):
        st.markdown("**Debug Mode Enabled** - This panel shows OpenAI request/response details.")

        # Show correlation ID
        if correlation_id:
            st.code(f"Correlation ID: {correlation_id}", language="text")
        else:
            st.warning("No correlation ID available (debug mode may have been off during request)")

        # Show error details if present
        if error:
            st.error("**Error Occurred:**")

            if isinstance(error, OpenAISchemaValidationError):
                st.markdown("**Type:** Schema Validation Error")
                st.markdown(f"**Schema:** {error.schema_name}")
                if error.decision_notes:
                    st.markdown(f"**Decision Notes:** {error.decision_notes}")
                if error.validation_errors:
                    st.markdown(f"**Validation Errors:** {len(error.validation_errors)}")
                    with st.expander("Show Validation Errors"):
                        st.json(error.validation_errors)
                if error.raw_output:
                    with st.expander("Show Raw Output"):
                        st.code(error.raw_output[:1000], language="json")  # Truncate to 1000 chars

            elif isinstance(error, OpenAITransportError):
                st.markdown("**Type:** Transport Error")
                if error.status_code:
                    st.markdown(f"**Status Code:** {error.status_code}")
                if error.retry_after:
                    st.markdown(f"**Retry After:** {error.retry_after}s")

            else:
                st.markdown(f"**Type:** {type(error).__name__}")
                st.markdown(f"**Message:** {str(error)}")

        # Load and show debug context from files
        if correlation_id:
            debug_context = get_debug_context(correlation_id)

            if debug_context:
                # Show request details
                if "request" in debug_context:
                    with st.expander("📤 Request Details"):
                        req = debug_context["request"]
                        st.markdown(f"**Model:** {req.get('model')}")
                        st.markdown(f"**Schema:** {req.get('schema_name')}")
                        st.markdown(f"**Timestamp:** {req.get('timestamp')}")

                        # Show messages (prompts)
                        if "request" in req and "messages" in req["request"]:
                            st.markdown("**Messages:**")
                            for msg in req["request"]["messages"]:
                                role = msg.get("role", "unknown")
                                content = msg.get("content", "")
                                st.text_area(
                                    f"Message ({role})",
                                    content[:500],  # Truncate to 500 chars
                                    height=150,
                                    key=f"msg_{role}_{correlation_id}"
                                )

                        # Download request JSON
                        request_json = json.dumps(req, indent=2)
                        st.download_button(
                            label="📥 Download Request JSON",
                            data=request_json,
                            file_name=f"request_{correlation_id}.json",
                            mime="application/json",
                            key=f"download_request_{correlation_id}"
                        )

                # Show response details
                if "response" in debug_context:
                    with st.expander("📥 Response Details"):
                        resp = debug_context["response"]
                        st.markdown(f"**Schema:** {resp.get('schema_name')}")
                        st.markdown(f"**Decision Notes:** {resp.get('decision_notes')}")
                        st.markdown(f"**Timestamp:** {resp.get('timestamp')}")

                        # Show metadata
                        if "response_metadata" in resp:
                            meta = resp["response_metadata"]
                            st.markdown("**Response Metadata:**")
                            st.json(meta)

                        # Show usage
                        if "usage" in resp:
                            usage = resp["usage"]
                            st.markdown("**Token Usage:**")
                            st.json(usage)

                        # Show refusal if present
                        if "refusal" in resp:
                            st.error(f"**Refusal:** {resp['refusal']}")

                        # Show output
                        if "output" in resp:
                            output = resp["output"]
                            st.markdown("**Output:**")
                            st.markdown(f"- Parsed: {output.get('parsed_present')}")
                            st.markdown(f"- Type: {output.get('parsed_type')}")
                            if output.get("text"):
                                st.text_area(
                                    "Output Text (truncated)",
                                    output["text"],
                                    height=150,
                                    key=f"output_{correlation_id}"
                                )

                        # Show error if present
                        if "error" in resp:
                            err = resp["error"]
                            st.error(f"**Error:** {err.get('type')} - {err.get('message')}")

                        # Download response JSON
                        response_json = json.dumps(resp, indent=2)
                        st.download_button(
                            label="📥 Download Response JSON",
                            data=response_json,
                            file_name=f"response_{correlation_id}.json",
                            mime="application/json",
                            key=f"download_response_{correlation_id}"
                        )

                # Show notes
                if "notes" in debug_context:
                    with st.expander("📝 Decision Notes"):
                        notes = debug_context["notes"]
                        st.json(notes)

            else:
                st.info(
                    "No debug files found. Set `CQC_OPENAI_DEBUG_SAVE_DIR` environment variable to save debug files.")

        # Add instructions
        st.markdown("---")
        st.markdown("""
        **Debug Mode Configuration:**
        - `CQC_OPENAI_DEBUG=1` - Enable debug mode
        - `CQC_OPENAI_DEBUG_REDACT=1` - Redact sensitive data (default: enabled)
        - `CQC_OPENAI_DEBUG_SAVE_DIR=/path/to/dir` - Save debug files to directory
        """)
